import os
import tempfile
import unittest

from openpyxl import load_workbook

from excel_generator import generate_airflow_pattern_excel


class AirflowPatternExcelTest(unittest.TestCase):
    def test_writes_clean_airflow_summary_rows(self):
        criteria = (
            '1. 육안상 단일방향류가 형성되어야 함.\n'
            '2. 측정대상 크린장비 내부에 난류가 형성되는 구역이 없어야 함.'
        )
        names = [
            '무균시험실 BSC',
            '균주접종실 BSC',
            '미생물 시험실 C/B(852)',
            '미생물 시험실 C/B(853)',
            'PASS BOX (QHA-745)',
            'PASS BOX (QHA-744)',
            'PASS BOX (QHA-743)',
            'PASS BOX (QHA-742)',
        ]
        items = [
            {
                'name': name,
                'date': '2025.08.02',
                'criteria': criteria,
                'video_attached': '첨부',
                'judgment': '적합',
            }
            for name in names
        ]
        data = {
            'unknown': [{
                'semester': '2025 (하)',
                'date': '2025.08.02',
                'items': items,
            }]
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'airflow.xlsx')
            generate_airflow_pattern_excel(data, path)
            workbook = load_workbook(path)

        sheet = workbook['AHU-unknown']
        self.assertEqual(
            [sheet.cell(4, col).value for col in range(1, 7)],
            ['NO', '측정대상', '측정기준', '동영상 첨부', '판정결과', '측정일자'],
        )
        self.assertEqual([sheet.cell(row, 2).value for row in range(5, 13)], names)
        for row in range(5, 13):
            self.assertEqual(sheet.cell(row, 3).value, criteria)
            self.assertEqual(sheet.cell(row, 4).value, '첨부')
            self.assertEqual(sheet.cell(row, 5).value, '적합')
            self.assertEqual(sheet.cell(row, 6).value, '2025 (하)')
            self.assertEqual(sheet.row_dimensions[row].height, 54)


if __name__ == '__main__':
    unittest.main()
