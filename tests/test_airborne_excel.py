import os
import tempfile
import unittest

from openpyxl import load_workbook

from excel_generator import generate_airborne_particle_excel


class AirborneParticleExcelTest(unittest.TestCase):
    def test_generates_workbook_with_numeric_and_unknown_ahu_keys(self):
        room = {
            'no_start': 1,
            'no_end': 1,
            'grade': 'B',
            'room_number': '2142',
            'room_name': '무균 실험실',
            'measurements': [{'point': 1, 'value_05': 121, 'value_50': 7}],
        }
        semester = {
            'semester': '2025 (하)',
            'date': '2025.08.02',
            'rooms': [room],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'airborne.xlsx')
            generate_airborne_particle_excel(
                {'unknown': [semester.copy()], '33': [semester.copy()]},
                path,
            )
            workbook = load_workbook(path, data_only=False)

        self.assertEqual(workbook.sheetnames[0], 'AHU-33 Data')
        self.assertIn('AHU-unknown Data', workbook.sheetnames)

    def test_writes_all_rooms_and_30_measurement_points(self):
        source_rooms = [
            ('B', '2142', '무균 실험실', [(121, 7), (194, 0), (676, 9), (576, 9), (107, 1), (121, 4)]),
            ('D', '2165', '균주접종실', [(5420, 400), (570, 20), (660, 50), (4000, 620), (690, 110), (600, 10)]),
            ('C', '2147', '탈의실', [(30760, 1340), (17690, 550)]),
            ('D', '2169', '복도', [(11140, 1670), (4120, 650), (4710, 590)]),
            ('D', '2145', '탈의실', [(54260, 1170)]),
            ('B', '2142', 'PASS BOX(QHA-745)', [(539, 4), (317, 0)]),
            ('B', '2142', 'PASS BOX(QHA-744)', [(90, 1), (114, 1)]),
            ('D', '2169', 'PASS BOX(QHA-743)', [(25830, 2240), (30810, 2960)]),
            ('D', '2169', 'PASS BOX(QHA-742)', [(60, 0), (70, 0)]),
            ('A', '2142', '무균시험실 BSC', [(3, 0), (1, 0)]),
            ('B', '2165', '균주접종실 BSC', [(4, 0), (0, 0)]),
        ]
        rooms = []
        no = 1
        for grade, room_number, name, values in source_rooms:
            measurements = [
                {'point': point, 'value_05': value_05, 'value_50': value_50}
                for point, (value_05, value_50) in enumerate(values, start=1)
            ]
            rooms.append({
                'no_start': no,
                'no_end': no + len(measurements) - 1,
                'grade': grade,
                'room_number': room_number,
                'room_name': name,
                'measurements': measurements,
            })
            no += len(measurements)

        data = {
            '37': [{
                'semester': '2025 (하)',
                'date': '2025.08.02',
                'rooms': rooms,
            }]
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'airborne.xlsx')
            generate_airborne_particle_excel(data, path)
            workbook = load_workbook(path, data_only=False)

        sheet = workbook['AHU-37 Data']
        self.assertEqual(sheet.max_row, 38)
        self.assertEqual([sheet.cell(row, 1).value for row in range(9, 39)], list(range(1, 31)))
        self.assertEqual(sheet['D21'].value, '탈의실')
        self.assertEqual(sheet['F21'].value, 30760)
        self.assertEqual(sheet['H21'].value, 1340)
        self.assertEqual(sheet['D37'].value, '균주접종실 BSC')
        for chart_sheet_name in ('AHU-37 0.5', 'AHU-37 5.0'):
            chart = workbook[chart_sheet_name]._charts[0]
            self.assertEqual(chart.x_axis.tickLblPos, 'low')
            self.assertEqual(chart.x_axis.txPr.bodyPr.rot, 0)
            self.assertEqual(chart.x_axis.txPr.p[0].pPr.defRPr.sz, 900)
            self.assertEqual(chart.anchor.ext.height, 18 * 360000)
            self.assertEqual(chart.anchor.ext.width, 49.5 * 360000)


if __name__ == '__main__':
    unittest.main()
