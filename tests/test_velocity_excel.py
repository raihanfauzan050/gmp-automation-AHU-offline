import os
import tempfile
import unittest

from openpyxl import load_workbook

from excel_generator import generate_air_velocity_excel


class AirVelocityExcelTest(unittest.TestCase):
    def test_chart_uses_readable_two_line_x_axis_labels(self):
        machines = []
        for index in range(8):
            machines.append({
                'no_start': index * 4 + 1,
                'no_end': index * 4 + 4,
                'grade': 'A' if index == 0 else 'B',
                'room_number': str(2142 + index),
                'machine_name': f'무균시험실 BSC\nBio Safety Cabinet-{index + 1}',
                'measurements': [
                    {'point': point, 'value': 0.4 + point / 100}
                    for point in range(1, 5)
                ],
            })
        data = {
            '33': [{
                'semester': '2025 (하)',
                'date': '2025.08.02',
                'machines': machines,
            }],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'velocity.xlsx')
            generate_air_velocity_excel(data, path)
            workbook = load_workbook(path, data_only=False)

        pivot = workbook['AHU-33 Pivot']
        self.assertEqual(pivot['D1'].value, '2025 (하) -')
        self.assertEqual(
            pivot['I2'].value,
            '=SUBSTITUTE(C2,CHAR(10)," ")&CHAR(10)&A2&" / "&B2',
        )
        chart = pivot._charts[0]
        self.assertEqual(chart.x_axis.tickLblPos, 'low')
        self.assertFalse(chart.x_axis.noMultiLvlLbl)
        self.assertEqual(chart.x_axis.txPr.bodyPr.rot, 0)
        self.assertEqual(chart.x_axis.txPr.p[0].pPr.defRPr.sz, 900)
        self.assertEqual(chart.anchor.ext.height, 18 * 360000)
        self.assertEqual(chart.anchor.ext.width, 36 * 360000)

    def test_creates_one_bar_series_per_semester_and_groups_by_equipment_id(self):
        def machine(name, value):
            return {
                'no_start': 1,
                'no_end': 4,
                'grade': 'B',
                'room_number': '2142',
                'machine_name': name,
                'measurements': [
                    {'point': point, 'value': value}
                    for point in range(1, 5)
                ],
            }

        data = {
            '33': [
                {
                    'semester': '2025 (하)',
                    'date': '2025.08.02',
                    'machines': [
                        machine('PASS BOX (QHA-744)', 0.40),
                        machine('PASS BOX (QHA-745)', 0.41),
                    ],
                },
                {
                    'semester': '2026 (하)',
                    'date': '2026.08.02',
                    'machines': [
                        machine('PASS BOX QHA 744', 0.42),
                        machine('PASS BOX QHA-745', 0.43),
                    ],
                },
            ],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, 'velocity.xlsx')
            generate_air_velocity_excel(data, path)
            workbook = load_workbook(path, data_only=False)

        pivot = workbook['AHU-33 Pivot']
        self.assertEqual(pivot.max_row, 3)
        self.assertEqual(pivot['D1'].value, '2026 (하) -')
        self.assertEqual(pivot['E1'].value, '2025 (하) -')
        self.assertEqual(pivot['D2'].value, "='AHU-33 Table'!$E$5")
        self.assertEqual(pivot['E2'].value, "='AHU-33 Table'!$E$7")
        self.assertEqual(len(pivot._charts[0]._charts[0].series), 2)


if __name__ == '__main__':
    unittest.main()
