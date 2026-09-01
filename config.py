"""
GMP Automation System - Configuration
All constants, alert limits, and test parameters
"""

import os

# =============================================================================
# DIRECTORIES
# =============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'outputs')

# =============================================================================
# OFFLINE OCR BACKEND
# =============================================================================
# The local DeepSeek-OCR server runs on this port outside Docker. Docker
# overrides it with the internal `ocr` service address.
DEEPSEEK_OCR_ENDPOINT = os.environ.get('DEEPSEEK_OCR_ENDPOINT', 'http://127.0.0.1:8000')

# =============================================================================
# A. AIRBORNE PARTICLE TEST
# =============================================================================
AIRBORNE_PARTICLE = {
    'name': 'Airborne Particle Test',
    'korean_title': '부유입자 측정 기록서',
    'excel_filename': 'Airborne_Particle_Test_Result_and_Graph.xlsx',
    'alert_limits': {
        '0.5': {'A': 23, 'B': 627, 'C': 23402, 'D': 141390},
        '5.0': {'A': 4, 'B': 13, 'C': 1540, 'D': 8183},
    },
    'action_limits': {
        '0.5': {'A': 46, 'B': 1254, 'C': 46804, 'D': 282780},
        '5.0': {'A': 8, 'B': 26, 'C': 2900, 'D': 16366},
    },
    'frequency': {
        'A': 6,   # months (6 = twice/year: Feb & Aug)
        'B': 6,
        'C': 12,  # months (12 = once/year: Aug only)
        'D': 12,
    },
    'sheets_per_ahu': 4,  # Data, Table, 0.5, 5.0
    'chart_y_max_05': 3000000,
    'chart_y_max_50': 30000,
}

# =============================================================================
# B. AIR VELOCITY TEST
# =============================================================================
AIR_VELOCITY = {
    'name': 'Air Velocity Test',
    'korean_title': '풍속 측정 기록지',
    'excel_filename': 'Air_Velocity_Test_Result_and_Graph.xlsx',
    'alert_limits': {
        'low': 0.38,   # below this = fail
        'high': 0.52,  # above this = fail
    },
    'action_limits': {
        'low': 0.36,
        'high': 0.54,
    },
    'frequency': {
        'A': 6, 'B': 6, 'C': 12, 'D': 12,
    },
    'sheets_per_ahu': 3,  # Data, Table, Pivot
    'chart_y_max': 0.8,
}

# =============================================================================
# C. AIR CHANGE RATE (ACH) TEST
# =============================================================================
AIR_CHANGE_RATE = {
    'name': 'Air Change Rate Test',
    'korean_title': '환기 횟수 측정 결과 기록서',
    'excel_filename': 'Air_Change_Rate_Test_Result_and_Graph.xlsx',
    'alert_limits': {
        'B': 52,  # below this = fail
        'C': 22,
        'D': 12,
    },
    'action_limits': {
        'B': 50,
        'C': 20,
        'D': 10,
    },
    'frequency': {
        'B': 6, 'C': 12, 'D': 12,
    },
    'sheets_per_ahu': 3,  # Data, Table, Pivot
    'chart_y_max': 100,
}

# =============================================================================
# D. HEPA FILTER TEST
# =============================================================================
HEPA_FILTER = {
    'name': 'HEPA Filter Test',
    'korean_title': 'HEPA FILTER 성능 검사 집계표',
    'excel_filename': 'HEPA_Filter_Test_Result_and_Graph.xlsx',
    'alert_limit': 0.0001,  # 0.01% as decimal
    'sheets_per_ahu': 3,  # Data, Table, Pivot
    'chart_y_max': 0.0002,  # 0.02% minimum chart maximum
}

# =============================================================================
# E. AIRFLOW PATTERN TEST
# =============================================================================
AIRFLOW_PATTERN = {
    'name': 'Airflow Pattern Test',
    'korean_title': '기류패턴시험 기록서',
    'excel_filename': 'Airflow_Pattern_Test_Result_and_Graph.xlsx',
    'pass_value': '적합',
    'sheets_per_ahu': 1,  # Only one sheet per AHU
}

# =============================================================================
# SEMESTER MAPPING
# =============================================================================
def get_semester_label(date_str):
    """Convert date string like '2025.08.02' to semester label like '2025 (하)'"""
    parts = date_str.replace('~', '').strip().split('.')
    year = parts[0].strip()
    month = int(parts[1].strip())
    if month == 8 or month > 6:
        return f"{year} (하)"
    else:
        return f"{year} (상)"

def semester_sort_key(label):
    """Sort key for semester labels. More recent = smaller key (goes to top)."""
    # e.g., "2025 (하)" -> (-2025, -2), "2025 (상)" -> (-2025, -1)
    year = int(label[:4])
    sem = -2 if '하' in label else -1
    return (-year, sem)

# =============================================================================
# EXCEL STYLING
# =============================================================================
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ALERT_FILL_RED = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16)
DATA_FONT = Font(size=11)

CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# TEST TYPE REGISTRY
# =============================================================================
TEST_TYPES = {
    'airborne_particle': AIRBORNE_PARTICLE,
    'air_velocity': AIR_VELOCITY,
    'air_change_rate': AIR_CHANGE_RATE,
    'hepa_filter': HEPA_FILTER,
    'airflow_pattern': AIRFLOW_PATTERN,
}
